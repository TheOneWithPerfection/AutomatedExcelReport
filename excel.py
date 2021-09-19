import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.chart import BarChart, Reference
import string

excel_file = pd.read_excel('supermarket_sales.xlsx')
excel_file[['Gender', 'Product line', 'Total']]

report_table = excel_file.pivot_table(index='Gender',
                                      columns='Product line',
                                      values='Total',
                                      aggfunc='sum').round(0)

report_table.to_excel('report_2021.xlsx',
                      sheet_name='Report',
                      startrow=4)

wb = load_workbook('report_2021.xlsx')
sheet = wb['Report']
# cell references (original spreadsheet) 
min_column = wb.active.min_column
max_column = wb.active.max_column
min_row = wb.active.min_row
max_row = wb.active.max_row

#
wb = load_workbook('report_2021.xlsx')
sheet = wb['Report']
# barchart
barchart = BarChart() #barchart = BarChart() initializes a barchart variable from the Barchart class.
#locate data and categories
data = Reference(sheet,
                 min_col=min_column+1,
                 max_col=max_column,
                 min_row=min_row,
                 max_row=max_row) #including headers
categories = Reference(sheet,
                       min_col=min_column,
                       max_col=min_column,
                       min_row=min_row+1,
                       max_row=max_row) #not including headers
#data and categories are variables that represent where that information is located. We’re using the column and row references we defined above to automate this. 

#We use add_data and set_categories to add the necessary data to the barchart. Inside add_data I’m adding the titles_from_data=True because I included the headers for data.
barchart.add_data(data, titles_from_data=True)
barchart.set_categories(categories)
#location chart
sheet.add_chart(barchart, "B12")
#We use sheet.add_chart to specify what we want to add to the “Report” sheet and in which cell we want to add it. 
barchart.title = 'Sales by Product line'
barchart.style = 5 #choose the chart style
#We can modify the default title and chart style using barchart.title and barchart.style
wb.save('report_2021.xlsx')
#We save all the changes with wb.save()

import string
alphabet = list(string.ascii_uppercase)
excel_alphabet = alphabet[0:max_column] 
print(excel_alphabet)
#If we print this we’ll obtain a list from A to G.
#This happens because first, we created an alphabet list from A to Z, 
#but then we took a slice [0:max_column] to match the length of this list (7) with the first 7 letters of the alphabet (A-G).
#Python lists start on 0, so A=0, B=1, C=2, and so on. Also, the [a:b] slice notation takes b-a elements (starting with “a” and ending with “b-1”)

wb = load_workbook('report_2021.xlsx')
sheet = wb['Report']
# sum in columns B-G
#for i in excel_alphabet loops through all the active columns, 
#but then we excluded the A column with if i!='A' because the A column doesn’t contain numeric data.
for i in excel_alphabet:
    if i!='A':
        sheet[f'{i}{max_row+1}'] = f'=SUM({i}{min_row+1}:{i}{max_row})'
        #sheet[f'{i}{max_row+1}'] = f'=SUM({i}{min_row+1}:{i}{max_row}' is the same as writing sheet['B7'] = '=SUM(B5:B6)' but now we do that for columns A to G.
        sheet[f'{i}{max_row+1}'].style = 'Currency'
        #sheet[f'{i}{max_row+1}'].style = 'Currency' gives the currency style to cells below the maximum row.
# adding total label
sheet[f'{excel_alphabet[0]}{max_row+1}'] = 'Total'
#We add the ‘Total’ label to the A column below the maximum row withsheet[f'{excel_alphabet[0]}{max_row+1}'] = 'Total'.
wb.save('report_2021.xlsx')

wb = load_workbook('report_2021.xlsx')
sheet = wb['Report']
sheet['A1'] = 'Sales Report'
sheet['A2'] = '2021'
sheet['A1'].font = Font('Arial', bold=True, size=20)
sheet['A2'].font = Font('Arial', bold=True, size=10)
wb.save('report_2021.xlsx')